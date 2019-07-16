from __future__ import print_function
from openpyxl import load_workbook
import datetime
import sys
import argparse
import os
from argparse import RawTextHelpFormatter
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")


def convertTimeToDecimalMinutes(durationSeq):
    """
    Converts a datetime variable into a number of minutes (decimal)

    :param durationSeq: datetime variable to convert
    :return: decimal number of minutes
    """

    durationSeqHour = durationSeq.hour
    durationSeqMinute = durationSeq.minute
    durationSeqSec = durationSeq.second
    time = durationSeqHour * 60 + durationSeqMinute + durationSeqSec / 60.0
    return time


def extractWhoAndWhat(actionWho, actionWhat, numSeq):

    returnString = ""

    listOfWho = actionWho.split(";")
    for who in listOfWho:
        returnString += who + " is " + str(numSeq) + "\n"

    if actionWhat == None:
        pass
    else:
        returnString += listOfWho[0] + " -> leg : " + actionWhat + "\n"

    return returnString


def generateDigitalBoard(celebrationName, cueSheet):
    """
    Generates the digital board file

    """

    fileArdoise = celebrationName + "_ARDOISE"

    ws = loadCueSheetFile(cueSheet)
    # getting the number of lines
    max_row = ws.max_row

    # creating an output file
    outputArdoise = open(fileArdoise, 'w')
	
    # processing the file
    print('\nDigital board: processing {} lines from the file.\n\n'.format(max_row), end='')
    for i in range (3, max_row):
            numSeq = ws['A' + str(i)].value
            timeSeq = ws['C' + str(i)].value
            actionSummary = ws['F' + str(i)].value
            print('.',end='')
            sys.stdout.flush()
            if numSeq == None:
                    pass
            else:
                    timeSeqBis = timeSeq.strftime('%-H|%-M|%-S')
                    if actionSummary == None:
                        outputArdoise.write(str(numSeq) + "|" + str(timeSeqBis) + "|\n")
                    else:
                        outputArdoise.write(str(numSeq) + "|" + str(timeSeqBis) + "|" + str(actionSummary) + "\n")

    # closing files
    outputArdoise.close()


def generatePlantUML(filename, options):
    if is_tool("plantuml"):
        # PlantUML is installed, let's process
        if options == "png":
            print("export to png")
            os.system('plantuml ' + '-t' + options + ' ' + filename)
        elif options == "svg":
            print("export to svg")
            os.system('plantuml ' + '-t' + options + ' ' + filename)
        else:
            print("export to png, because default")
            os.system('plantuml ' + filename)
        
    else:
        # PlantUML is not installed, so display a message instead
        print("The plantUML file has been generated, now render it with PlantUML.\n" +
            "If you do not have PlantUML on your computer, you can render online\n" +
            "by pasting the code on https://www.planttext.com")


def generateTimeDiagram(celebrationName, cueSheet, exportOptions):
    """
    Generates the time diagram txt file

    """

    fileDiagram = celebrationName + "_DIAGRAM"

    ws = loadCueSheetFile(cueSheet)
    # getting the number of lines
    max_row = ws.max_row

    # creating an output file
    outputDiagram = open(fileDiagram, 'w')

    # processing the file
    print('\nTime diagram: processing {} lines from the file.\n\n'.format(max_row), end='')

    outputDiagram.write(timeDiagramHeader(celebrationName))

    firstDiagramLine = True

    for i in range (15, max_row):
            numSeq = ws['A' + str(i)].value
            actionWho = ws['G' + str(i)].value
            actionWhat = ws['H' + str(i)].value

            if actionWho == None:
                pass
            else:
                if firstDiagramLine:
                    outputDiagram.write("\' sequence " + str(numSeq) + "\n@0\n")
                    firstDiagramLine = False

                    outputDiagram.write(extractWhoAndWhat(actionWho, actionWhat, numSeq))

                else:
                    durationSeq = ws['B' + str(i-1)].value                       
                    outputDiagram.write("\' sequence " + str(numSeq) + "\n@+" + str(convertTimeToDecimalMinutes(durationSeq)) + "\n")
                    outputDiagram.write(extractWhoAndWhat(actionWho, actionWhat, numSeq))

                    listOfPreviousWho = ws['G' + str(i-1)].value.split(";")
                    listOfCurrentWho = actionWho.split(";")
                    for who in listOfPreviousWho:
                        if who not in listOfCurrentWho:
                            outputDiagram.write(who + " is {hidden}\n")
                            
    outputDiagram.write(timeDiagramFooter())

    # closing files
    outputDiagram.close()

    # generate plantUML if installed
    generatePlantUML(fileDiagram, exportOptions)


def is_tool(name):
    """
    Check whether `name` is on PATH.
    """

    from distutils.spawn import find_executable

    return find_executable(name) is not None


def loadCueSheetFile(cueSheet):
    """
    Loads the cue sheet XLS file

    :param file: file
    :return:
    """

    print('\nAnalyzing "{}" XLS file...'.format(cueSheet), end='\n')

    wb = load_workbook(cueSheet, keep_vba=True, data_only=True)

    # selecting the relevant tab
    ws = wb['ArdNum']
    return ws


def print_list_of_puml_exports():
    """
    Prints the list of plant UML export types

    :return:
    """
    return('[EXPORT] defines the type of export you require.\n' +
           '* png (default)\n' +
           '* svg')


def print_list_of_types():
    """
    Prints the list of output types

    :return:
    """
    return('[TYPE] defines the type of output you require.\n' +
            '* db: outputs the digital board file (ardoise numerique)\n' +
            '* td: outputs the time diagram\n' +
            '* by default, both will be generated')


def timeDiagramFooter():
    """
    Returns the footer for the PlantUML time diagram

    :return: footer in string
    """

    footer = "\' end of time diagram\n@enduml"
    return footer


def timeDiagramHeader(title):
    """
    Returns the header for the PlantUML time diagram

    :param title: title to display on the time diagram
    :return: header in string
    """

    header = "@startuml\n\n"
    header += "skinparam monochrome true\nTitle " + title + "\n\n"
    header += "concise \"CHANTS\" as song\n"
    header += "concise \"MGR\" as mgr\n"
    header += "concise \"PADRE\" as padre\n"
    header += "concise \"VIDEO\" as vid\n"
    header += "concise \"INTERV\" as int\n"
    header += "concise \"TXT/PRAY\" as txt\n"
    header += "concise \"ACTORS\" as act\n"
    header += "concise \" \" as leg\n"
    header += "scale 5 as 140 pixels\n\n"

    return header


def main(argv):

	# Get parameters from command line
	parser = argparse.ArgumentParser(description='Extracts info from FRAT cue sheets',
                formatter_class=RawTextHelpFormatter)
	parser.add_argument('-f', '--file', help='input XLS file', required=True)
	parser.add_argument('-o', '--output', help='output name', required=True)
        parser.add_argument('-t', '--type', help=print_list_of_types())
        parser.add_argument('-e', '--export', help=print_list_of_puml_exports())
	args = parser.parse_args()

	cueSheetPath = Path(args.file)
        celebrationName = args.output

        if cueSheetPath.is_file():
            if args.type == "db":
                # digital board (ardoise numerique)
                cueSheet = args.file
                generateDigitalBoard(celebrationName, cueSheet)
            elif args.type == "td":
                # time diagram
                cueSheet = args.file
                exportOptions = args.export
                generateTimeDiagram(celebrationName, cueSheet, exportOptions)
            else:
                # both digital board and time diagram
                cueSheet = args.file
                exportOptions = args.export
                generateDigitalBoard(celebrationName, cueSheet)
                generateTimeDiagram(celebrationName, cueSheet, exportOptions)
        else:
            print("Input file does not exist!")


if __name__ == '__main__':
	sys.exit(main(sys.argv))
